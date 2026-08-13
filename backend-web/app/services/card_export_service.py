"""
卡券数据导出服务

功能：
1. 导出卡券为 Excel（.xlsx），包含「卡券」Sheet
2. 可选导出「卡券商品关联」Sheet（卡券 ↔ 商品关联信息）
3. 关联信息与闲鱼账号关联（通过 xy_catalog_items：item_id → account_pk），
   支持按所选闲鱼账号过滤只导出这些账号的关联信息

设计要点：
- 所有字段以字符串形式写入（复用账号导出的 _to_str 思路），避免 Excel 自动转数字
- 与账号导出（account_export_service）保持相同的 Sheet 风格，便于互操作
"""
from __future__ import annotations

import io
import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from common.models.card import Card
from common.models.card_item_relation import CardItemRelation
from common.models.xy_account import XYAccount
from common.models.xy_catalog_item import XYCatalogItem
from common.utils.time_utils import get_beijing_now


def _to_str(value) -> str:
    """将任意值转为字符串，避免 Excel 自动转数字。None 转为空字符串。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _write_sheet(wb: Workbook, sheet_name: str, headers: list[str], rows: list[list]) -> None:
    """向工作簿写入一个 Sheet，所有单元格设为文本格式。"""
    ws = wb.create_sheet(title=sheet_name)
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_to_str(value))
            cell.number_format = "@"
    # 自动调整列宽（简单估算）
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, min(len(rows) + 2, 50)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value or ""
            max_len = max(max_len, min(len(str(cell_val)), 60))
        ws.column_dimensions[col_letter].width = max_len + 2


class CardExportService:
    """卡券数据导出服务"""

    # 「卡券」Sheet 表头
    CARD_HEADERS = [
        "卡券名称", "类型", "描述", "启用", "延迟秒数",
        "多规格", "规格名", "规格值",
        "对接价格", "是否可对接", "手续费支付方", "最低售价", "对接可见性",
        "API配置", "文本内容", "数据内容", "图片URL", "多图片URL",
    ]

    # 「卡券商品关联」Sheet 表头
    RELATION_HEADERS = ["卡券名称", "商品ID", "来源"]

    def __init__(self, session: AsyncSession):
        self.session = session

    async def export_cards(
        self,
        owner_id: int | None,
        include_relations: bool = False,
        account_ids: list[str] | None = None,
    ) -> io.BytesIO:
        """导出卡券数据为 Excel。

        Args:
            owner_id: 所有者ID（None=管理员导出全部）
            include_relations: 是否包含「卡券商品关联」Sheet
            account_ids: 关联信息按账号过滤的账号ID列表；None/空 = 不过滤（全部账号）

        Returns:
            Excel 文件的 BytesIO 对象
        """
        # 1. 查询卡券
        stmt = select(Card).order_by(Card.id.asc())
        if owner_id is not None:
            stmt = stmt.where(Card.user_id == owner_id)
        result = await self.session.execute(stmt)
        cards = list(result.scalars().all())
        card_id_to_name = {c.id: c.name for c in cards}

        # 2. 生成 Excel
        wb = Workbook()
        wb.remove(wb.active)
        self._write_cards_sheet(wb, cards)

        if include_relations:
            relations = await self._get_relations(cards, account_ids)
            self._write_relations_sheet(wb, relations, card_id_to_name)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ==================== 数据查询 ====================

    async def _get_relations(
        self,
        cards: list[Card],
        account_ids: list[str] | None,
    ) -> list[CardItemRelation]:
        """查询卡券商品关联。

        当指定了账号过滤时，只返回 item_id 属于所选账号商品的关联记录
        （商品归属通过 xy_catalog_items.account_pk → xy_accounts 判断）。
        """
        card_ids = [c.id for c in cards]
        if not card_ids:
            return []

        stmt = select(CardItemRelation).where(CardItemRelation.card_id.in_(card_ids))
        result = await self.session.execute(stmt)
        relations = list(result.scalars().all())

        # 账号过滤：解析所选账号的商品 item_id 集合
        selected_item_ids: set[str] | None = None
        if account_ids:
            acc_result = await self.session.execute(
                select(XYAccount.id).where(XYAccount.account_id.in_(account_ids))
            )
            account_pks = [row[0] for row in acc_result.all()]
            if account_pks:
                item_result = await self.session.execute(
                    select(XYCatalogItem.item_id).where(
                        XYCatalogItem.account_pk.in_(account_pks)
                    )
                )
                selected_item_ids = {row[0] for row in item_result.all()}
            else:
                # 所选账号不存在于当前库：过滤后为空
                selected_item_ids = set()

        if selected_item_ids is not None:
            relations = [r for r in relations if r.item_id in selected_item_ids]

        return relations

    # ==================== Sheet 写入 ====================

    def _write_cards_sheet(self, wb: Workbook, cards: list[Card]) -> None:
        rows = []
        for card in cards:
            rows.append([
                card.name, card.type, card.description, card.enabled, card.delay_seconds,
                card.is_multi_spec, card.spec_name, card.spec_value,
                card.price, card.is_dockable, card.fee_payer, card.min_price,
                card.dock_visibility,
                card.api_config, card.text_content, card.data_content,
                card.image_url, card.image_urls,
            ])
        _write_sheet(wb, "卡券", self.CARD_HEADERS, rows)

    def _write_relations_sheet(
        self,
        wb: Workbook,
        relations: list[CardItemRelation],
        card_id_to_name: dict[int, str],
    ) -> None:
        rows = []
        for rel in relations:
            card_name = card_id_to_name.get(rel.card_id, "")
            rows.append([card_name, rel.item_id, rel.source])
        _write_sheet(wb, "卡券商品关联", self.RELATION_HEADERS, rows)

    @staticmethod
    def build_filename() -> str:
        """生成导出文件名（北京时间）。"""
        now = get_beijing_now()
        return f"cards_export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
