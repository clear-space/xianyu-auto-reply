"""
卡券数据导入服务

功能：
1. 解析卡券导出/账号导出的 Excel 文件（openpyxl）
2. 预览：解析文件内容，返回卡券列表（含是否已存在）与关联列表
   （含账号归属、是否在账号过滤范围内），供前端勾选导入
3. 「卡券」Sheet：按 卡券名称+规格值 upsert（已存在则更新，不存在则新建），
   支持按预览行号选择性导入
4. 可选「卡券商品关联」Sheet：卡券↔商品关联导入（已存在跳过），
   支持按预览行号选择性导入
5. 关联信息与闲鱼账号关联，支持按所选闲鱼账号过滤：
   只导入 item_id 归属所选账号商品（目标库 xy_catalog_items）的关联行，
   无法判定归属的关联行跳过并计数提示
"""
from __future__ import annotations

from io import BytesIO
from typing import Any

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from common.models.card import Card
from common.models.card_item_relation import CardItemRelation
from common.models.xy_account import XYAccount
from common.models.xy_catalog_item import XYCatalogItem

# 三态哨兵：区分「未传 match_owner_id」（=按 owner 范围检测）与「显式传 None」（=全库检测）
_MATCH_OWNER_UNSET = object()


def _parse_bool(value: Any) -> bool:
    """解析布尔值：'是'/True/'true'/'1' → True，其余 → False"""
    if value is None:
        return False
    v = str(value).strip().lower()
    return v in ("是", "true", "1", "yes")


def _parse_int(value: Any, default: int = 0) -> int:
    """解析整数，失败返回默认值"""
    if value is None or str(value).strip() == "":
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def _parse_str(value: Any) -> str:
    """解析字符串，None → 空字符串"""
    if value is None:
        return ""
    return str(value).strip()


def _read_sheet_rows(wb, sheet_name: str) -> list[dict[str, str]]:
    """读取 Sheet 为字典列表（表头作为 key），保留行序号于 _row_index 字段"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    result = []
    for row_idx, row in enumerate(rows[1:]):
        row_dict = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            val = row[i] if i < len(row) else None
            row_dict[header] = str(val).strip() if val is not None else ""
        # 跳过全空行
        if any(v for v in row_dict.values()):
            row_dict["_row_index"] = row_idx  # 0-based，对应数据区行号
            result.append(row_dict)
    return result


class CardImportService:
    """卡券数据导入服务"""

    def __init__(
        self,
        session: AsyncSession,
        owner_id: int,
        match_owner_id: object = _MATCH_OWNER_UNSET,
    ):
        """初始化导入服务。

        Args:
            session: 数据库会话
            owner_id: 新建卡券/关联归属的用户 ID（当前登录用户）
            match_owner_id: 重复卡券检测的用户范围
                - 不传（默认）: 仅在 owner_id 范围内按名称检测（普通用户语义）
                - None: 全库按名称检测，不限归属（管理员语义）
                - int: 在该用户范围内检测
        """
        self.session = session
        self.owner_id = owner_id
        # 重复检测范围：None 表示不限制归属（全库按 名称+规格值 匹配）
        if match_owner_id is _MATCH_OWNER_UNSET:
            self.match_owner_id: int | None = owner_id
        else:
            self.match_owner_id = match_owner_id  # type: ignore[assignment]
        # 导入过程中的映射缓存：卡券名称|规格值 → 卡券ID
        self._card_name_spec_to_id: dict[str, int] = {}

    async def import_cards(
        self,
        file_content: bytes,
        include_relations: bool = False,
        account_ids: list[str] | None = None,
        card_indexes: list[int] | None = None,
        relation_indexes: list[int] | None = None,
        duplicate_mode: str = "overwrite",
    ) -> dict:
        """导入卡券数据。

        Args:
            file_content: Excel 文件内容
            include_relations: 是否导入「卡券商品关联」Sheet
            account_ids: 关联信息按账号过滤的账号ID列表；None/空 = 不过滤（全部导入）
            card_indexes: 只导入指定行号的卡券（None = 全部，行号来自预览结果）
            relation_indexes: 只导入指定行号的关联（None = 全部）
            duplicate_mode: 重复卡券（同名同规格已存在）处理方式
                - overwrite: 覆盖更新已存在的卡券（默认）
                - skip: 跳过已存在的卡券，不更新

        Returns:
            {success, message, data: {card_inserted, card_updated,
             card_skipped_duplicate, relation_inserted,
             relation_skipped_exists, relation_skipped_account, errors}}
        """
        if duplicate_mode not in ("overwrite", "skip"):
            duplicate_mode = "overwrite"

        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            return {
                "success": False,
                "message": f"Excel 文件解析失败: {str(e)}",
                "data": None,
            }

        stats = {
            "card_inserted": 0,
            "card_updated": 0,
            "card_skipped_duplicate": 0,
            "relation_inserted": 0,
            "relation_skipped_exists": 0,
            "relation_skipped_account": 0,
            "errors": [],
        }

        try:
            await self._import_cards(wb, stats, card_indexes, duplicate_mode)
            if include_relations:
                await self._import_relations(
                    wb, account_ids, stats, relation_indexes
                )
        except Exception as e:
            logger.opt(exception=e).error("卡券导入失败")
            await self.session.rollback()
            return {
                "success": False,
                "message": f"导入失败: {str(e)[:300]}",
                "data": stats,
            }

        return {"success": True, "message": "导入完成", "data": stats}

    async def preview_cards(
        self,
        file_content: bytes,
        include_relations: bool = False,
        account_ids: list[str] | None = None,
    ) -> dict:
        """解析 Excel 文件，返回预览数据供前端勾选导入。

        Args:
            file_content: Excel 文件内容
            include_relations: 是否解析「卡券商品关联」Sheet
            account_ids: 关联信息按账号过滤的账号ID列表；None/空 = 不过滤

        Returns:
            {success, message, data: {
                cards: [{index, name, type, spec_name, spec_value, description, exists}],
                relations: [{index, card_name, item_id, source, account_ids, in_scope}],
            }}
        """
        try:
            wb = load_workbook(BytesIO(file_content), read_only=True, data_only=True)
        except Exception as e:
            return {
                "success": False,
                "message": f"Excel 文件解析失败: {str(e)}",
                "data": None,
            }

        try:
            cards, relations = await self._build_preview(
                wb, include_relations, account_ids
            )
        except Exception as e:
            logger.opt(exception=e).error("卡券导入预览失败")
            return {
                "success": False,
                "message": f"预览解析失败: {str(e)[:300]}",
                "data": None,
            }

        return {
            "success": True,
            "data": {"cards": cards, "relations": relations},
        }

    def _apply_match_scope(self, stmt):
        """按 match_owner_id 追加卡券归属过滤；None = 全库不过滤（管理员语义）。"""
        if self.match_owner_id is not None:
            stmt = stmt.where(Card.user_id == self.match_owner_id)
        return stmt

    async def _build_preview(
        self,
        wb,
        include_relations: bool,
        account_ids: list[str] | None,
    ) -> tuple[list[dict], list[dict]]:
        """构建预览数据：卡券列表（含是否存在）与关联列表（含账号归属与过滤范围）。"""
        # 1. 卡券：解析 + 批量判断是否已存在
        card_rows = _read_sheet_rows(wb, "卡券")
        existing_keys: set[str] = set()
        existing_stmt = self._apply_match_scope(
            select(Card.name, Card.spec_value)
        )
        result = await self.session.execute(existing_stmt)
        for name, spec_val in result.all():
            existing_keys.add(f"{name}|{spec_val or ''}")

        cards = []
        for row in card_rows:
            name = _parse_str(row.get("卡券名称"))
            if not name:
                continue
            spec_value = _parse_str(row.get("规格值"))
            cards.append({
                "index": int(row.get("_row_index", 0)),
                "name": name,
                "type": _parse_str(row.get("类型")) or "text",
                "spec_name": _parse_str(row.get("规格名")),
                "spec_value": spec_value,
                "description": _parse_str(row.get("描述")),
                "exists": f"{name}|{spec_value}" in existing_keys,
            })

        # 2. 关联：解析 + 账号归属 + 过滤范围
        relations: list[dict] = []
        if include_relations:
            # item_id -> 归属账号集合
            item_accounts: dict[str, list[str]] = {}
            join_stmt = (
                select(XYCatalogItem.item_id, XYAccount.account_id)
                .join(XYAccount, XYAccount.id == XYCatalogItem.account_pk)
            )
            result = await self.session.execute(join_stmt)
            for item_id, acc_id in result.all():
                item_accounts.setdefault(item_id, []).append(acc_id)

            allowed_item_ids = await self._resolve_allowed_item_ids(account_ids)

            for row in _read_sheet_rows(wb, "卡券商品关联"):
                card_name = _parse_str(row.get("卡券名称"))
                item_id = _parse_str(row.get("商品ID"))
                if not card_name or not item_id:
                    continue
                acc_ids = sorted(set(item_accounts.get(item_id, [])))
                relations.append({
                    "index": int(row.get("_row_index", 0)),
                    "card_name": card_name,
                    "item_id": item_id,
                    "source": _parse_str(row.get("来源")) or "own",
                    "account_ids": acc_ids,
                    "in_scope": allowed_item_ids is None or item_id in allowed_item_ids,
                })

        return cards, relations

    async def _resolve_allowed_item_ids(
        self, account_ids: list[str] | None
    ) -> set[str] | None:
        """解析账号过滤对应的商品 item_id 集合；None = 不过滤。"""
        if not account_ids:
            return None
        acc_result = await self.session.execute(
            select(XYAccount.id).where(XYAccount.account_id.in_(account_ids))
        )
        account_pks = [row[0] for row in acc_result.all()]
        if not account_pks:
            return set()
        item_result = await self.session.execute(
            select(XYCatalogItem.item_id).where(
                XYCatalogItem.account_pk.in_(account_pks)
            )
        )
        return {row[0] for row in item_result.all()}

    # ==================== 卡券导入 ====================

    async def _import_cards(
        self,
        wb,
        stats: dict,
        card_indexes: list[int] | None,
        duplicate_mode: str,
    ) -> None:
        """按 名称+规格值 upsert 卡券（可指定行号范围）。

        duplicate_mode=overwrite 时已存在的卡券被更新；
        duplicate_mode=skip 时已存在的卡券跳过（不更新）。
        """
        index_set = set(card_indexes) if card_indexes is not None else None
        rows = _read_sheet_rows(wb, "卡券")
        for row in rows:
            if index_set is not None and int(row.get("_row_index", -1)) not in index_set:
                continue
            name = _parse_str(row.get("卡券名称"))
            if not name:
                continue
            card_type = _parse_str(row.get("类型")) or "text"
            spec_value = _parse_str(row.get("规格值"))

            stmt = self._apply_match_scope(
                select(Card).where(Card.name == name)
            )
            if spec_value:
                stmt = stmt.where(Card.spec_value == spec_value)
            else:
                stmt = stmt.where((Card.spec_value.is_(None)) | (Card.spec_value == ""))

            result = await self.session.execute(stmt)
            existing = result.scalars().first()

            if existing:
                if duplicate_mode == "skip":
                    stats["card_skipped_duplicate"] += 1
                    self._card_name_spec_to_id[f"{name}|{spec_value}"] = existing.id
                    continue
                self._apply_card_row(existing, row, card_type, spec_value)
                self.session.add(existing)
                stats["card_updated"] += 1
            else:
                card = Card(
                    user_id=self.owner_id,
                    name=name,
                    type=card_type,
                    description=_parse_str(row.get("描述")) or None,
                    enabled=_parse_bool(row.get("启用")),
                    delay_seconds=_parse_int(row.get("延迟秒数"), 0),
                    price=_parse_str(row.get("对接价格")) or None,
                    is_dockable=_parse_bool(row.get("是否可对接")),
                    fee_payer=_parse_str(row.get("手续费支付方")) or None,
                    min_price=_parse_str(row.get("最低售价")) or None,
                    dock_visibility=_parse_str(row.get("对接可见性")) or None,
                    is_multi_spec=_parse_bool(row.get("多规格")),
                    spec_name=_parse_str(row.get("规格名")) or None,
                    spec_value=spec_value or None,
                    api_config=_parse_str(row.get("API配置")) or None,
                    text_content=_parse_str(row.get("文本内容")) or None,
                    data_content=_parse_str(row.get("数据内容")) or None,
                    image_url=_parse_str(row.get("图片URL")) or None,
                    image_urls=_parse_str(row.get("多图片URL")) or None,
                )
                self.session.add(card)
                await self.session.flush()
                stats["card_inserted"] += 1

            self._card_name_spec_to_id[f"{name}|{spec_value}"] = (
                existing.id if existing else card.id
            )

        await self.session.commit()

        # 补充映射：查询匹配范围内的既有卡券，覆盖 Sheet 之外的卡券（关联 Sheet 可能引用）
        stmt = self._apply_match_scope(
            select(Card.id, Card.name, Card.spec_value)
        )
        result = await self.session.execute(stmt)
        for card_id, card_name, spec_val in result.all():
            key = f"{card_name}|{spec_val or ''}"
            self._card_name_spec_to_id.setdefault(key, card_id)

    def _apply_card_row(self, card: Card, row: dict, card_type: str, spec_value: str) -> None:
        """将 Sheet 行数据更新到已存在的卡券对象。"""
        card.type = card_type
        card.description = _parse_str(row.get("描述")) or card.description
        card.enabled = _parse_bool(row.get("启用"))
        card.delay_seconds = _parse_int(row.get("延迟秒数"), 0)
        card.price = _parse_str(row.get("对接价格")) or card.price
        card.is_dockable = _parse_bool(row.get("是否可对接"))
        card.fee_payer = _parse_str(row.get("手续费支付方")) or card.fee_payer
        card.min_price = _parse_str(row.get("最低售价")) or card.min_price
        card.dock_visibility = _parse_str(row.get("对接可见性")) or card.dock_visibility
        card.is_multi_spec = _parse_bool(row.get("多规格"))
        card.spec_name = _parse_str(row.get("规格名")) or card.spec_name
        card.spec_value = spec_value or card.spec_value
        card.api_config = _parse_str(row.get("API配置")) or card.api_config
        card.text_content = _parse_str(row.get("文本内容")) or card.text_content
        card.data_content = _parse_str(row.get("数据内容")) or card.data_content
        card.image_url = _parse_str(row.get("图片URL")) or card.image_url
        card.image_urls = _parse_str(row.get("多图片URL")) or card.image_urls

    # ==================== 关联导入 ====================

    async def _import_relations(
        self,
        wb,
        account_ids: list[str] | None,
        stats: dict,
        relation_indexes: list[int] | None,
    ) -> None:
        """导入卡券商品关联（按账号过滤，可指定行号范围）。"""
        rows = _read_sheet_rows(wb, "卡券商品关联")
        if not rows:
            return

        index_set = set(relation_indexes) if relation_indexes is not None else None
        # 账号过滤：解析所选账号在目标库的商品 item_id 集合
        allowed_item_ids = await self._resolve_allowed_item_ids(account_ids)

        # 本次导入已处理的 (card_id, item_id) 集合，防止 Excel 内重复行触发唯一键冲突
        seen_relation_keys: set[tuple[int, str]] = set()

        for row in rows:
            if index_set is not None and int(row.get("_row_index", -1)) not in index_set:
                continue
            card_name = _parse_str(row.get("卡券名称"))
            item_id = _parse_str(row.get("商品ID"))
            if not card_name or not item_id:
                continue

            # 账号过滤：item_id 归属所选账号的商品才导入
            if allowed_item_ids is not None and item_id not in allowed_item_ids:
                stats["relation_skipped_account"] += 1
                continue

            # 查找卡券ID（先按名称精确匹配，与账号导入保持一致的宽松策略）
            card_id = self._card_name_spec_to_id.get(f"{card_name}|")
            if not card_id:
                for key, cid in self._card_name_spec_to_id.items():
                    if key.startswith(f"{card_name}|"):
                        card_id = cid
                        break
            if not card_id:
                continue

            relation_key = (card_id, item_id)
            if relation_key in seen_relation_keys:
                stats["relation_skipped_exists"] += 1
                continue

            # 已存在则跳过。
            # 注意：唯一键 uk_card_item_dock 是 (card_id, item_id, dock_record_id)，
            # 不包含 user_id —— 去重检查必须与唯一键对齐（不限定 user_id），
            # 否则跨用户同卡同商品的既有关联会导致 INSERT 撞唯一键（历史 bug）。
            stmt = select(CardItemRelation).where(
                CardItemRelation.card_id == card_id,
                CardItemRelation.item_id == item_id,
            )
            result = await self.session.execute(stmt)
            if result.scalars().first():
                stats["relation_skipped_exists"] += 1
                seen_relation_keys.add(relation_key)
                continue
            seen_relation_keys.add(relation_key)

            source = _parse_str(row.get("来源")) or "own"
            rel = CardItemRelation(
                user_id=self.owner_id,
                card_id=card_id,
                item_id=item_id,
                source=source,
            )
            self.session.add(rel)
            stats["relation_inserted"] += 1

        await self.session.commit()
